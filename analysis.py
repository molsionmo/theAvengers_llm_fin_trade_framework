import pandas as pd
import numpy as np
import argparse
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


def calculate_financial_metrics(csv_file, risk_free_rate=0.02):
    """
    从CSV文件中读取数据，计算累计回报率、年收益率、夏普比率和最大回撤。
    :param csv_file: CSV文件路径
    :param risk_free_rate: 无风险利率，默认为0.02
    :return: 包含计算结果的DataFrame
    """
    # 读取CSV文件
    data = pd.read_csv(csv_file, index_col=0, parse_dates=True)

    # 计算每日收益率
    daily_returns = data['Close'].pct_change()

    # 初始化结果列表
    cumulative_returns_list = []
    annual_return_list = []
    sharpe_ratio_list = []
    max_drawdown_list = []

    for i in range(len(data)):
        # 截取当前时间点之前的数据
        sub_data = data.iloc[:i + 1]
        sub_daily_returns = daily_returns.iloc[:i + 1]

        # 计算累计回报率
        cumulative_returns = (1 + sub_daily_returns).cumprod()
        cumulative_returns_list.append(cumulative_returns[-1])

        # 计算年收益率
        if len(sub_data) > 1:
            years = (sub_data.index[-1] - sub_data.index[0]).days / 365
            annual_return = (cumulative_returns[-1]) ** (1 / years) - 1
        else:
            annual_return = np.nan
        annual_return_list.append(annual_return)

        # 计算夏普比率
        if len(sub_daily_returns) > 1:
            daily_std = sub_daily_returns.std()
            sharpe_ratio = (annual_return - risk_free_rate) / (daily_std * np.sqrt(252))
        else:
            sharpe_ratio = np.nan
        sharpe_ratio_list.append(sharpe_ratio)


    # 整理结果
    result_df = pd.DataFrame({
        '累计回报率': cumulative_returns_list,
        '年收益率': annual_return_list,
        '夏普比率': sharpe_ratio_list,
    }, index=data.index)

    return result_df


def adjust_column_widths(file_path):
    """
    调整Excel文件的列宽以适应内容
    :param file_path: Excel文件路径
    """
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="计算股票的金融指标并保存为Excel文件")
    parser.add_argument("--csv-file", type=str, required=True, help="输入的CSV文件路径")
    parser.add_argument("--output", type=str, required=True, help="输出的Excel文件路径")
    parser.add_argument("--risk-free-rate", type=float, default=0.02, help="无风险利率，默认为0.02")

    args = parser.parse_args()

    csv_file = args.csv_file
    output_file = args.output
    risk_free_rate = args.risk_free_rate

    # 计算金融指标
    result_df = calculate_financial_metrics(csv_file, risk_free_rate)

    # 保存结果到Excel文件
    result_df.to_excel(output_file, index=True)

    # 调整列宽
    adjust_column_widths(output_file)

    print(f"计算结果已保存到 {output_file}")